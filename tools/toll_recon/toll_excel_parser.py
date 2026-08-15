#!/usr/bin/env python
"""Read the FASTag wallet statements the banks send, one folder per truck.

TWO BANKS, TWO LAYOUTS. 79 of the 84 workbooks are LIVQUIK "Fastag Wallet
Transaction Report" sheets: a block of account headers, then a row starting
"SN", then one row per transaction with the lorry number in `truckNo` and the
plaza in `otherPartyName`. The other 5 are IDFC statements with a completely
different shape — the lorry number appears once in the page header, the plaza is
buried in a sentence ("Issuer Debit Transaction for toll fare - 378005 -
Nazirakhat Toll Plaza"), and every toll debit is shadowed by a Credit of the
same amount, which is the wallet topping itself back up and NOT a toll.

THE DATES ARE THE DANGEROUS PART. LIVQUIK writes txnDate in day-first order, but
only some of the workbooks store it as text; in the rest Excel has already
parsed it, and it parsed day-first text as month-first. So 1,440 of the 1,587
date cells come out of openpyxl as a real datetime pointing at the wrong day —
a July statement whose rows claim to be in February. Nothing about the cell
says it is wrong. Read naively, 41% of these tolls land on a date that never
happened, which breaks the five-minute duplicate check and attaches the cost to
whatever trip was running months earlier.

What makes it recoverable is that every statement declares its own period. So
the period is the referee: a date is accepted only if it falls inside the window
the file itself claims to cover, and day/month are swapped when — and only when
— that is what puts it there. Rows that cannot be placed inside their own
statement's window are refused rather than guessed at; a toll on an invented
date is worse than a toll not imported.

(Where day equals month the swap is a no-op and the two readings agree, so those
147 rows are not ambiguous, merely symmetric.)

Output is JSON for the importer, which does the de-duplication against tolls the
GTROPY API already pulled, the trip matching and the posting. This file only
reads paper.
"""
from __future__ import annotations

import argparse
import datetime as dt
import glob
import io
import json
import os
import re
import sys
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

ROOTS = {
    "Jaiswal Toll": ("JAISWAL ENTERPRISE", "FASTag Wallet: Jaiswal Enterprise"),
    "Prasad Toll":  ("PRASAD TRANSPORT",   "FASTag Wallet: Prasad Transport"),
}

# "Issuer Debit Transaction for toll fare - 378005 - Nazirakhat Toll Plaza"
IDFC_PLAZA = re.compile(r"toll fare\s*-\s*(\d+)\s*-\s*(.+)$", re.I)


def norm_reg(txt: str | None) -> str | None:
    """The same normalisation the database applies, character for character.

    Postgres: upper(regexp_replace(txt, '[^A-Za-z0-9]', '', 'g')). Deliberately
    NOT str.isalnum(), which is unicode-aware and would keep a Devanagari digit
    that Postgres would have stripped — the two sides must agree exactly or a
    lookup by vehicle_no_norm silently misses.
    """
    if not txt:
        return None
    return "".join(c for c in str(txt).upper() if ("A" <= c <= "Z") or ("0" <= c <= "9")) or None


def _cells(path: str) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    finally:
        wb.close()


def _window(rows: list[tuple]) -> tuple[dt.date | None, dt.date | None]:
    """The period the statement says it covers — the referee for every date."""
    frm = to = None
    for row in rows[:30]:
        j = " ".join(str(c) for c in row if c is not None)
        m = re.search(r"From Date\s*:\s*(\d{4}-\d{2}-\d{2})", j)
        if m:
            frm = dt.date.fromisoformat(m.group(1))
        m = re.search(r"To Date\s*:\s*(\d{4}-\d{2}-\d{2})", j)
        if m:
            to = dt.date.fromisoformat(m.group(1))
        m = re.search(r"Statement Duration", j)
        if m:
            got = re.findall(r"(\d{2}\s+\w{3}\s+\d{4})", j)
            if len(got) == 2:
                frm = dt.datetime.strptime(got[0], "%d %b %Y").date()
                to = dt.datetime.strptime(got[1], "%d %b %Y").date()
    return frm, to


def _place(value, frm: dt.date | None, to: dt.date | None) -> tuple[dt.datetime | None, str | None]:
    """Resolve one txnDate against the window its own statement declares."""
    if value is None:
        return None, "no date"

    cands: list[dt.datetime] = []
    if isinstance(value, dt.datetime):
        cands.append(value)
        try:                       # Excel read this pump's day-first text as month-first
            cands.append(value.replace(day=value.month, month=value.day))
        except ValueError:
            pass
    else:
        s = str(value).strip()
        for fmt in ("%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%Y-%m-%d %H:%M:%S",
                    "%d %b %y %I:%M %p", "%d %b %Y %I:%M %p"):
            try:
                cands.append(dt.datetime.strptime(s, fmt))
            except ValueError:
                pass
        if not cands:
            return None, f"unparseable date {s!r}"

    if not (frm and to):
        return cands[0], None      # no window to judge against; take it as written

    inside = [c for c in cands if frm <= c.date() <= to]
    if not inside:
        return None, (f"{cands[0]:%Y-%m-%d} is outside the statement's own "
                      f"period {frm}..{to} under every reading")
    # Symmetric readings (day == month) collapse to one value, so this is a
    # choice only in name.
    if len({c for c in inside}) > 1:
        return None, "two different dates both fall inside the statement period"
    return inside[0], None


def parse_livquik(path: str, rows: list[tuple], group: str, company: str) -> tuple[list[dict], Counter]:
    stat = Counter()
    hdr_i = next((i for i, r in enumerate(rows)
                  if r and r[0] is not None and str(r[0]).strip() == "SN"), None)
    if hdr_i is None:
        return [], stat
    hdr = [str(c).strip() if c is not None else "" for c in rows[hdr_i]]
    col = {name: i for i, name in enumerate(hdr)}
    frm, to = _window(rows)

    out = []
    for r in rows[hdr_i + 1:]:
        if not r or r[0] is None:
            continue
        get = lambda k: r[col[k]] if k in col and col[k] < len(r) else None
        if str(get("type") or "").strip().upper() != "DEBIT":
            stat["not a debit"] += 1
            continue
        when, why = _place(get("txnDate"), frm, to)
        if when is None:
            stat[f"refused: {why}"] += 1
            continue
        amt = get("amount")
        try:
            amt = float(amt)
        except (TypeError, ValueError):
            stat["refused: unreadable amount"] += 1
            continue
        if amt <= 0:
            stat["refused: non-positive amount"] += 1
            continue
        out.append({
            "source_file": os.path.basename(path), "group": group, "company_hint": company,
            "bank": "LIVQUIK",
            "ext_txn_id": str(get("externalId")).strip() if get("externalId") else None,
            "txn_ref": str(get("txnRef")).strip() if get("txnRef") else None,
            "vehicle_raw": str(get("truckNo") or "").strip() or None,
            "vehicle_norm": norm_reg(get("truckNo")),
            "txn_datetime": when.isoformat(sep=" "),
            "plaza_name": str(get("otherPartyName") or "").strip() or None,
            "toll_id": str(get("tollId")).strip() if get("tollId") else None,
            "lane": str(get("lane")).strip() if get("lane") else None,
            "direction": str(get("direction")).strip() if get("direction") else None,
            "tag_id": str(get("kitNo")).strip() if get("kitNo") else None,
            "amount": round(amt, 2),
        })
        stat["read"] += 1
    return out, stat


def parse_idfc(path: str, rows: list[tuple], group: str, company: str) -> tuple[list[dict], Counter]:
    """IDFC: lorry number in the page header, plaza inside a sentence, and every
    debit shadowed by a same-amount Credit that is the wallet refilling itself."""
    stat = Counter()
    hdr_i = next((i for i, r in enumerate(rows)
                  if r and r[0] is not None and str(r[0]).strip() == "Transaction Time"), None)
    if hdr_i is None:
        return [], stat

    veh = None
    for r in rows[:hdr_i]:
        if r and r[0] and "truck number" in str(r[0]).strip().lower():
            veh = str(r[1]).strip() if len(r) > 1 and r[1] else None
    frm, to = _window(rows)

    out = []
    for r in rows[hdr_i + 1:]:
        if not r or r[0] is None:
            continue
        nature = str(r[1] or "").strip().upper()
        if nature != "DEBIT":
            stat["wallet top-up, not a toll" if nature == "CREDIT" else "not a debit"] += 1
            continue
        when, why = _place(r[0], frm, to)
        if when is None:
            stat[f"refused: {why}"] += 1
            continue
        try:
            amt = float(r[2])
        except (TypeError, ValueError):
            stat["refused: unreadable amount"] += 1
            continue
        if amt <= 0:
            stat["refused: non-positive amount"] += 1
            continue
        desc = str(r[3] or "")
        m = IDFC_PLAZA.search(desc)
        out.append({
            "source_file": os.path.basename(path), "group": group, "company_hint": company,
            "bank": str(r[7]).strip() if len(r) > 7 and r[7] else "IDFC",
            "ext_txn_id": str(r[4]).strip() if len(r) > 4 and r[4] else None,
            "txn_ref": None,
            "vehicle_raw": veh, "vehicle_norm": norm_reg(veh),
            "txn_datetime": when.isoformat(sep=" "),
            "plaza_name": (m.group(2).strip() if m else desc.strip()) or None,
            "toll_id": m.group(1) if m else None,
            "lane": None, "direction": None, "tag_id": None,
            "amount": round(amt, 2),
        })
        stat["read"] += 1
    return out, stat


def run(roots: list[str], out_json: str | None) -> None:
    files = []
    for root in roots:
        label = os.path.basename(root.rstrip("/\\"))
        company, _wallet = ROOTS.get(label, ("UNKNOWN", None))
        for p in sorted(glob.glob(os.path.join(root, "**", "*.xlsx"), recursive=True)):
            files.append((p, label, company))

    all_rows: list[dict] = []
    stats = Counter()
    per_bank = Counter()
    skipped_files = []

    for path, group, company in files:
        rows = _cells(path)
        got, st = parse_livquik(path, rows, group, company)
        if not got and not st:
            got, st = parse_idfc(path, rows, group, company)
        if not got and not st:
            skipped_files.append(path)
        all_rows += got
        stats += st
        for g in got:
            per_bank[g["bank"]] += 1

    print(f"  workbooks      : {len(files)}")
    print(f"  toll debits read: {len(all_rows)}")
    print(f"  by bank        : {dict(per_bank)}")
    if all_rows:
        print(f"  date range     : {min(r['txn_datetime'] for r in all_rows)[:10]}"
              f" .. {max(r['txn_datetime'] for r in all_rows)[:10]}")
        print(f"  value          : {sum(r['amount'] for r in all_rows):,.2f}")
        print(f"  vehicles       : {len({r['vehicle_norm'] for r in all_rows})}")
        print(f"  with ext id    : {sum(1 for r in all_rows if r['ext_txn_id'])}")

    print("\n  rows not taken:")
    for k, n in stats.most_common():
        if k != "read":
            print(f"   {n:>5}  {k}")
    if skipped_files:
        print(f"\n  UNRECOGNISED LAYOUT ({len(skipped_files)}):")
        for p in skipped_files:
            print(f"    {p}")

    if out_json:
        os.makedirs(os.path.dirname(out_json) or ".", exist_ok=True)
        with open(out_json, "w", encoding="utf-8") as fh:
            json.dump(all_rows, fh, indent=1)
        print(f"\n  wrote {len(all_rows)} rows -> {out_json}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", action="append", required=True,
                    help="a toll folder (repeatable)")
    ap.add_argument("--json", help="write parsed rows here")
    a = ap.parse_args()
    run(a.root, a.json)
